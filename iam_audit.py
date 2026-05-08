"""List all IAM users with their policies, categorized by admin-level access."""

import argparse
import boto3
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

ADMIN_POLICIES = {
    "arn:aws:iam::aws:policy/AdministratorAccess",
    "arn:aws:iam::aws:policy/IAMFullAccess",
    "arn:aws:iam::aws:policy/PowerUserAccess",
}

POWER_POLICIES = {
    "arn:aws:iam::aws:policy/PowerUserAccess",
    "arn:aws:iam::aws:policy/job-function/NetworkAdministrator",
    "arn:aws:iam::aws:policy/job-function/DatabaseAdministrator",
}

COLORS = {
    "red_dark":    "C00000",
    "red_light":   "FFCCCC",
    "orange_dark": "E36C09",
    "orange_light":"FCE4D6",
    "blue_dark":   "2F5496",
    "blue_light":  "D9E1F2",
    "green_dark":  "375623",
    "green_light": "E2EFDA",
    "grey_dark":   "595959",
    "grey_light":  "F2F2F2",
    "white":       "FFFFFF",
    "header_bg":   "1F3864",
}

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def get_iam_client(role_arn=None):
    if role_arn:
        sts = boto3.client("sts")
        creds = sts.assume_role(RoleArn=role_arn, RoleSessionName="iam-audit")["Credentials"]
        return boto3.client(
            "iam",
            aws_access_key_id=creds["AccessKeyId"],
            aws_secret_access_key=creds["SecretAccessKey"],
            aws_session_token=creds["SessionToken"],
        )
    return boto3.client("iam")


iam = None


def is_admin_inline(policy_doc):
    for stmt in policy_doc.get("Statement", []):
        actions = stmt.get("Action", [])
        resources = stmt.get("Resource", [])
        if isinstance(actions, str):
            actions = [actions]
        if isinstance(resources, str):
            resources = [resources]
        if stmt.get("Effect") == "Allow" and "*" in actions and "*" in resources:
            return True
    return False


def has_iam_star(policy_doc):
    """Check if policy grants iam:* (privilege escalation risk)."""
    for stmt in policy_doc.get("Statement", []):
        actions = stmt.get("Action", [])
        if isinstance(actions, str):
            actions = [actions]
        if stmt.get("Effect") == "Allow" and any(
            a in ("iam:*", "*") for a in actions
        ):
            return True
    return False


def days_since(dt):
    if dt is None:
        return None
    now = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return (now - dt).days


# Patterns for classifying standard users
BASELINE_POLICIES = {"IAMUserChangePassword", "MFAaccess", "IAMReadOnlyAccess", "EnforceMFA", "MFA",
                     "AWSRevokeOlderSessions"}
READONLY_KEYWORDS = {"readonly", "readonlyaccess", "read-only", "viewonly", "describe", "list", "get"}
SERVICE_ACCOUNT_NAMES = {"api-gateway-access", "ses-smtp", "airflow", "homework-sheets-ml",
                         "insight-extractor-worker", "lokis3bucket", "cloud_cleaner",
                         "ECS-user", "API_Access", "Platform", "Bachpan-Buddy", "s3-refshift-eks"}

# Map of AWS service prefixes to friendly names
SERVICE_PREFIX_MAP = {
    "redshift": "Redshift", "quicksight": "QuickSight", "polly": "Polly",
    "s3": "S3", "ec2": "EC2", "ecs": "ECS", "lambda": "Lambda",
    "dynamodb": "DynamoDB", "apigateway": "API Gateway", "api-gateway": "API Gateway",
    "glue": "Glue", "athena": "Athena", "cloudwatch": "CloudWatch",
    "eventbridge": "EventBridge", "ssm": "SSM", "codebuild": "CodeBuild",
    "codepipeline": "CodePipeline", "grafana": "Grafana", "ses": "SES",
    "secretsmanager": "Secrets Manager", "eks": "EKS", "ecr": "ECR",
    "vpc": "VPC", "billing": "Billing", "servicequotas": "Service Quotas",
}


def classify_standard_user(username, policies):
    """Classify a standard user into a sub-category based on policy patterns."""
    # Check service account by name
    if username in SERVICE_ACCOUNT_NAMES or not any(c in username for c in ["@", "."]):
        # Non-email usernames are likely service accounts (heuristic)
        meaningful = [p for p in policies if not any(b in p for b in BASELINE_POLICIES)]
        if meaningful:
            return "Service Account"

    # Strip baseline policies to analyze real access
    meaningful = []
    for p in policies:
        pname = p.split("] ")[1] if "] " in p else p
        if pname.split(" ⚠")[0].strip() not in BASELINE_POLICIES:
            meaningful.append(pname.lower())

    if not meaningful:
        return "Minimal Access"

    # Check if all meaningful policies are read-only
    is_readonly = all(any(kw in p for kw in READONLY_KEYWORDS) for p in meaningful)
    if is_readonly:
        return "ReadOnly"

    # Detect which AWS services are referenced
    detected_services = set()
    for p in meaningful:
        for prefix, svc in SERVICE_PREFIX_MAP.items():
            if prefix in p:
                detected_services.add(svc)

    has_write = any("full" in p or "write" in p or "crud" in p or "admin" in p for p in meaningful)

    if len(detected_services) == 1:
        svc = next(iter(detected_services))
        return f"Service-Specific ({svc})" if has_write else f"Service-Specific ({svc}, ReadOnly)"
    elif len(detected_services) >= 3 and has_write:
        return "Multi-Service Write"
    elif len(detected_services) >= 2:
        return "Multi-Service" + (" Write" if has_write else " ReadOnly")
    elif has_write:
        return "Single-Service Write"
    else:
        return "Limited Access"


def get_user_detail(username):
    """Return enriched user detail: policies, admin flag, category, MFA, last activity."""
    policies = []
    is_admin = False
    is_power = False
    has_iam_escalation = False

    # Attached managed policies
    for page in iam.get_paginator("list_attached_user_policies").paginate(UserName=username):
        for p in page["AttachedPolicies"]:
            arn = p["PolicyArn"]
            policies.append(f"[Direct] {p['PolicyName']}")
            if arn in ADMIN_POLICIES - POWER_POLICIES:
                is_admin = True
            if arn in POWER_POLICIES:
                is_power = True

    # Inline policies
    for page in iam.get_paginator("list_user_policies").paginate(UserName=username):
        for name in page["PolicyNames"]:
            doc = iam.get_user_policy(UserName=username, PolicyName=name)["PolicyDocument"]
            admin_flag = is_admin_inline(doc)
            iam_flag = has_iam_star(doc)
            tag = " ⚠ ADMIN-LEVEL" if admin_flag else (" ⚠ IAM-ESCALATION" if iam_flag else "")
            policies.append(f"[Inline] {name}{tag}")
            if admin_flag:
                is_admin = True
            if iam_flag:
                has_iam_escalation = True

    # Group policies
    for page in iam.get_paginator("list_groups_for_user").paginate(UserName=username):
        for g in page["Groups"]:
            gname = g["GroupName"]
            for gpage in iam.get_paginator("list_attached_group_policies").paginate(GroupName=gname):
                for p in gpage["AttachedPolicies"]:
                    arn = p["PolicyArn"]
                    policies.append(f"[Group:{gname}] {p['PolicyName']}")
                    if arn in ADMIN_POLICIES - POWER_POLICIES:
                        is_admin = True
                    if arn in POWER_POLICIES:
                        is_power = True
            for gpage in iam.get_paginator("list_group_policies").paginate(GroupName=gname):
                for name in gpage["PolicyNames"]:
                    doc = iam.get_group_policy(GroupName=gname, PolicyName=name)["PolicyDocument"]
                    admin_flag = is_admin_inline(doc)
                    iam_flag = has_iam_star(doc)
                    tag = " ⚠ ADMIN-LEVEL" if admin_flag else (" ⚠ IAM-ESCALATION" if iam_flag else "")
                    policies.append(f"[Group:{gname}/Inline] {name}{tag}")
                    if admin_flag:
                        is_admin = True
                    if iam_flag:
                        has_iam_escalation = True

    # Determine category
    if is_admin:
        category = "Admin"
        sub_category = ""
    elif has_iam_escalation:
        category = "Privileged"
        sub_category = ""
    elif is_power:
        category = "Power User"
        sub_category = ""
    elif not policies:
        category = "No Policies"
        sub_category = ""
    else:
        category = "Standard"
        sub_category = classify_standard_user(username, policies)

    # MFA status
    try:
        mfa_devices = iam.list_mfa_devices(UserName=username)["MFADevices"]
        mfa_enabled = "Yes" if mfa_devices else "No"
    except Exception:
        mfa_enabled = "Unknown"

    # Last activity via access keys
    last_used = None
    try:
        keys = iam.list_access_keys(UserName=username)["AccessKeyMetadata"]
        for key in keys:
            lu = iam.get_access_key_last_used(AccessKeyId=key["AccessKeyId"])
            used_date = lu.get("AccessKeyLastUsed", {}).get("LastUsedDate")
            if used_date:
                if last_used is None or used_date > last_used:
                    last_used = used_date
    except Exception:
        pass

    return {
        "category": category,
        "sub_category": sub_category,
        "policies": policies,
        "mfa_enabled": mfa_enabled,
        "last_used": last_used,
        "days_inactive": days_since(last_used),
    }


# ── Styling helpers ────────────────────────────────────────────────────────────

def hdr_cell(ws, row, col, value, bg=COLORS["header_bg"], fg=COLORS["white"], size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=fg, size=size, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN_BORDER
    return c


def data_cell(ws, row, col, value, bg=COLORS["white"], bold=False, wrap=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=bold)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = THIN_BORDER
    return c


CATEGORY_COLORS = {
    "Admin":      (COLORS["red_dark"],    COLORS["red_light"]),
    "Privileged": (COLORS["orange_dark"], COLORS["orange_light"]),
    "Power User": (COLORS["orange_dark"], COLORS["orange_light"]),
    "Standard":   (COLORS["blue_dark"],   COLORS["blue_light"]),
    "No Policies":(COLORS["grey_dark"],   COLORS["grey_light"]),
}


def write_user_sheet(ws, users, title):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color=COLORS["white"], size=13, name="Arial")
    c.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "User Name", "Category", "Sub-Category", "Created Date",
               "MFA Enabled", "Days Inactive", "Policies", "Notes"]
    for col, h in enumerate(headers, 1):
        hdr_cell(ws, 2, col, h)
    ws.row_dimensions[2].height = 22

    for idx, u in enumerate(users, 1):
        row = idx + 2
        cat = u["category"]
        _, row_bg = CATEGORY_COLORS.get(cat, (COLORS["blue_dark"], COLORS["blue_light"]))
        alt_bg = "F7FBFF" if idx % 2 == 0 else row_bg

        policies_text = "\n".join(u["policies"]) if u["policies"] else "(no policies attached)"
        days = u["days_inactive"]
        inactive_str = f"{days} days" if days is not None else "Never used"

        # Flag notes
        notes = []
        if cat == "Admin":
            notes.append("⚠ Admin-level access – review immediately")
        if u["mfa_enabled"] == "No":
            notes.append("⚠ MFA not enabled")
        if days is not None and days > 90:
            notes.append(f"⚠ Inactive {days}+ days")
        if cat == "No Policies":
            notes.append("ℹ No policies – consider removing user")

        data_cell(ws, row, 1, idx, alt_bg, align="center")
        data_cell(ws, row, 2, u["name"], alt_bg, bold=(cat == "Admin"))
        # Category badge
        cat_fg, cat_bg = CATEGORY_COLORS.get(cat, (COLORS["blue_dark"], COLORS["blue_light"]))
        c = ws.cell(row=row, column=3, value=cat)
        c.font = Font(name="Arial", size=10, bold=True, color=cat_fg)
        c.fill = PatternFill("solid", fgColor=cat_bg)
        c.alignment = Alignment(horizontal="center", vertical="top")
        c.border = THIN_BORDER

        data_cell(ws, row, 4, u.get("sub_category", "") or "—", alt_bg, align="center")
        data_cell(ws, row, 5, u["created"], alt_bg, align="center")
        mfa_bg = COLORS["green_light"] if u["mfa_enabled"] == "Yes" else COLORS["red_light"]
        c = ws.cell(row=row, column=6, value=u["mfa_enabled"])
        c.font = Font(name="Arial", size=10, bold=True,
                      color=COLORS["green_dark"] if u["mfa_enabled"] == "Yes" else COLORS["red_dark"])
        c.fill = PatternFill("solid", fgColor=mfa_bg)
        c.alignment = Alignment(horizontal="center", vertical="top")
        c.border = THIN_BORDER

        inactive_bg = COLORS["red_light"] if (days is not None and days > 90) else alt_bg
        data_cell(ws, row, 7, inactive_str, inactive_bg, align="center")
        data_cell(ws, row, 8, policies_text, alt_bg, wrap=True)
        data_cell(ws, row, 9, "\n".join(notes) if notes else "—", alt_bg, wrap=True)

        line_count = max(policies_text.count("\n") + 1, len(notes) or 1)
        ws.row_dimensions[row].height = max(20, min(line_count * 15, 120))

    col_widths = [5, 24, 14, 22, 14, 12, 14, 58, 38]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Auto-filter on header row
    ws.auto_filter.ref = f"A2:I{len(users) + 2}"


def write_summary_sheet(ws, all_users):
    ws.sheet_view.showGridLines = False
    from collections import Counter

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"IAM User Audit – Summary Dashboard  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"
    c.font = Font(bold=True, color=COLORS["white"], size=13, name="Arial")
    c.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cat_counts = Counter(u["category"] for u in all_users)
    sub_counts = Counter(u.get("sub_category", "") for u in all_users if u["category"] == "Standard")
    mfa_off = sum(1 for u in all_users if u["mfa_enabled"] == "No")
    inactive_90 = sum(1 for u in all_users if u["days_inactive"] is not None and u["days_inactive"] > 90)
    total = len(all_users) or 1

    # KPI boxes
    kpis = [
        ("Total Users",   len(all_users),             COLORS["blue_dark"],   COLORS["blue_light"]),
        ("Admin Users",   cat_counts.get("Admin", 0), COLORS["red_dark"],    COLORS["red_light"]),
        ("MFA Disabled",  mfa_off,                    COLORS["orange_dark"], COLORS["orange_light"]),
        ("Inactive 90d+", inactive_90,                COLORS["orange_dark"], COLORS["orange_light"]),
        ("Privileged",    cat_counts.get("Privileged", 0) + cat_counts.get("Power User", 0),
                                                       COLORS["orange_dark"], COLORS["orange_light"]),
        ("No Policies",   cat_counts.get("No Policies", 0), COLORS["grey_dark"], COLORS["grey_light"]),
    ]
    for i, (label, val, fg, bg) in enumerate(kpis, 1):
        lc = ws.cell(row=3, column=i, value=label)
        lc.font = Font(bold=True, name="Arial", size=10, color=fg)
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = THIN_BORDER
        vc = ws.cell(row=4, column=i, value=val)
        vc.font = Font(bold=True, name="Arial", size=20, color=fg)
        vc.fill = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = THIN_BORDER
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 36
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ── Section 1: High-Level Category Breakdown ──
    row = 6
    ws.merge_cells(f"A{row}:F{row}")
    h = ws[f"A{row}"]
    h.value = "Section 1: High-Level Category Breakdown"
    h.font = Font(bold=True, color=COLORS["white"], size=11, name="Arial")
    h.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

    row += 1
    for col, hd in enumerate(["Category", "Count", "% of Total", "Risk Level", "Recommended Action"], 1):
        hdr_cell(ws, row, col, hd)

    risk_map = {
        "Admin":      ("🔴 Critical", "Review & limit immediately; enforce MFA; consider SSO roles"),
        "Privileged": ("🟠 High",     "Audit permissions; apply least-privilege; monitor CloudTrail"),
        "Power User": ("🟠 Medium",   "Validate necessity; scope down to service-specific access"),
        "Standard":   ("🟢 Low",      "See sub-category breakdown below for targeted actions"),
        "No Policies":("⚪ Info",     "Remove unused accounts or assign appropriate policies"),
    }
    for cat in ["Admin", "Privileged", "Power User", "Standard", "No Policies"]:
        row += 1
        cnt = cat_counts.get(cat, 0)
        risk, action = risk_map[cat]
        cat_bg = CATEGORY_COLORS.get(cat, (COLORS["blue_dark"], COLORS["blue_light"]))[1]
        bg = cat_bg if row % 2 == 0 else COLORS["white"]
        data_cell(ws, row, 1, cat, bg, bold=True)
        data_cell(ws, row, 2, cnt, bg, align="center")
        pc = ws.cell(row=row, column=3, value=cnt / total)
        pc.number_format = "0.0%"
        pc.font = Font(name="Arial", size=10)
        pc.fill = PatternFill("solid", fgColor=bg)
        pc.alignment = Alignment(horizontal="center", vertical="top")
        pc.border = THIN_BORDER
        data_cell(ws, row, 4, risk, bg)
        data_cell(ws, row, 5, action, bg, wrap=True)
        ws.row_dimensions[row].height = 22

    # ── Section 2: Standard User Sub-Category Breakdown ──
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    h = ws[f"A{row}"]
    h.value = "Section 2: Standard User Sub-Category Breakdown & Recommendations"
    h.font = Font(bold=True, color=COLORS["white"], size=11, name="Arial")
    h.fill = PatternFill("solid", fgColor=COLORS["blue_dark"])
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

    row += 1
    for col, hd in enumerate(["Sub-Category", "Count", "% of Std", "Risk Level", "Recommended Action", "Review Frequency"], 1):
        hdr_cell(ws, row, col, hd)

    std_total = cat_counts.get("Standard", 0) or 1
    sub_risk_map = {
        "Multi-Service Write":  ("🟠 Medium-High", "Broad write access across services. Validate each service is needed; use Access Analyzer to generate least-privilege policies; consider splitting into service-specific roles.", "Monthly"),
        "Service Account":      ("🟠 Medium",      "Programmatic access. Rotate access keys regularly; ensure no console access; restrict to exact API actions needed; tag for automated monitoring.", "Monthly"),
        "Multi-Service ReadOnly":("🟢 Low-Medium", "Read access across multiple services. Acceptable for analysts/auditors; verify no write actions are needed; check for data-sensitive services.", "Quarterly"),
        "Service-Specific":     ("🟢 Low",         "Scoped to a single service. Good least-privilege pattern; verify the access level (Full vs Read) matches the role; check for inactive accounts.", "Quarterly"),
        "ReadOnly":             ("🟢 Low",         "Read-only access. Lowest risk standard pattern; ensure MFA is enabled; routine review only.", "Semi-annually"),
        "Single-Service Write": ("🟡 Low-Medium",  "Write access to one service. Verify the service and actions are appropriate for the user's role; consider scoping to specific resources.", "Quarterly"),
        "Limited Access":       ("🟢 Low",         "Narrow permissions that don't fit other patterns. Review policies to confirm they match job function.", "Quarterly"),
        "Minimal Access":       ("⚪ Info",        "Only baseline IAM/MFA policies. Confirm if user still needs an account; may be candidates for removal if inactive.", "Semi-annually"),
    }

    # Sort sub-categories by count descending
    sorted_subs = sorted(sub_counts.items(), key=lambda x: -x[1])
    # Also include any sub-categories with 0 count from the map
    seen = {s for s, _ in sorted_subs}
    for s in sub_risk_map:
        if s not in seen:
            sorted_subs.append((s, 0))

    for sub, cnt in sorted_subs:
        if not sub:
            continue
        row += 1
        # Match sub-category to risk map; handle Service-Specific variants
        key = sub
        if sub.startswith("Service-Specific"):
            key = "Service-Specific"
        risk, action, freq = sub_risk_map.get(key, ("🟢 Low", "Review policies for appropriateness.", "Quarterly"))
        bg = COLORS["blue_light"] if row % 2 == 0 else COLORS["white"]
        data_cell(ws, row, 1, sub, bg, bold=True)
        data_cell(ws, row, 2, cnt, bg, align="center")
        pc = ws.cell(row=row, column=3, value=cnt / std_total)
        pc.number_format = "0.0%"
        pc.font = Font(name="Arial", size=10)
        pc.fill = PatternFill("solid", fgColor=bg)
        pc.alignment = Alignment(horizontal="center", vertical="top")
        pc.border = THIN_BORDER
        data_cell(ws, row, 4, risk, bg)
        data_cell(ws, row, 5, action, bg, wrap=True)
        data_cell(ws, row, 6, freq, bg, align="center")
        ws.row_dimensions[row].height = 44

    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 18

    # ── Section 3: High-Risk Users Quick Reference ──
    risky = [u for u in all_users if u["category"] in ("Admin", "Privileged", "Power User")]
    if risky:
        row += 2
        ws.merge_cells(f"A{row}:F{row}")
        h = ws[f"A{row}"]
        h.value = "Section 3: High-Risk Users – Quick Reference"
        h.font = Font(bold=True, color=COLORS["white"], size=11, name="Arial")
        h.fill = PatternFill("solid", fgColor=COLORS["red_dark"])
        h.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22

        row += 1
        for col, sh in enumerate(["User Name", "Category", "MFA", "Days Inactive", "Policy Count", "Notes"], 1):
            hdr_cell(ws, row, col, sh, bg=COLORS["red_dark"])

        for u in risky:
            row += 1
            notes = []
            if u["mfa_enabled"] == "No":
                notes.append("No MFA")
            if u["days_inactive"] is not None and u["days_inactive"] > 90:
                notes.append(f"Inactive {u['days_inactive']}d")
            days_str = f"{u['days_inactive']}d" if u["days_inactive"] is not None else "Never"
            bg = COLORS["red_light"] if row % 2 == 0 else COLORS["white"]
            data_cell(ws, row, 1, u["name"], bg, bold=True)
            data_cell(ws, row, 2, u["category"], bg)
            data_cell(ws, row, 3, u["mfa_enabled"], bg, align="center")
            data_cell(ws, row, 4, days_str, bg, align="center")
            data_cell(ws, row, 5, len(u["policies"]), bg, align="center")
            data_cell(ws, row, 6, "; ".join(notes) if notes else "—", bg)


def write_sub_category_sheet(ws, std_users):
    """Write standard users grouped by sub-category with per-group headers."""
    ws.sheet_view.showGridLines = False
    from collections import defaultdict

    groups = defaultdict(list)
    for u in std_users:
        # Normalize Service-Specific variants for grouping
        sub = u.get("sub_category", "") or "Uncategorized"
        groups[sub].append(u)

    # Sort groups: Multi-Service Write first (highest risk), then others
    group_order = ["Multi-Service Write", "Service Account", "Multi-Service ReadOnly",
                   "Single-Service Write", "Limited Access"]
    sorted_keys = []
    for k in group_order:
        if k in groups:
            sorted_keys.append(k)
    for k in sorted(groups.keys()):
        if k not in sorted_keys:
            sorted_keys.append(k)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Standard Users – Sub-Category Detail"
    c.font = Font(bold=True, color=COLORS["white"], size=13, name="Arial")
    c.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    row = 3
    for sub in sorted_keys:
        users = groups[sub]
        # Group header
        ws.merge_cells(f"A{row}:F{row}")
        h = ws[f"A{row}"]
        h.value = f"{sub}  ({len(users)} users)"
        h.font = Font(bold=True, color=COLORS["white"], size=11, name="Arial")
        h.fill = PatternFill("solid", fgColor=COLORS["blue_dark"])
        h.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1

        for col, hd in enumerate(["User Name", "Created", "MFA", "Days Inactive", "Policies", "Notes"], 1):
            hdr_cell(ws, row, col, hd)
        row += 1

        for u in users:
            days = u["days_inactive"]
            days_str = f"{days}d" if days is not None else "Never"
            policies_text = "\n".join(u["policies"]) if u["policies"] else "—"
            notes = []
            if u["mfa_enabled"] == "No":
                notes.append("No MFA")
            if days is not None and days > 90:
                notes.append(f"Inactive {days}d")
            bg = COLORS["blue_light"] if row % 2 == 0 else COLORS["white"]
            data_cell(ws, row, 1, u["name"], bg, bold=True)
            data_cell(ws, row, 2, u["created"], bg, align="center")
            data_cell(ws, row, 3, u["mfa_enabled"], bg, align="center")
            data_cell(ws, row, 4, days_str, bg, align="center")
            data_cell(ws, row, 5, policies_text, bg, wrap=True)
            data_cell(ws, row, 6, "; ".join(notes) if notes else "—", bg, wrap=True)
            line_count = policies_text.count("\n") + 1
            ws.row_dimensions[row].height = max(20, min(line_count * 15, 100))
            row += 1
        row += 1  # blank row between groups

    for i, w in enumerate([24, 12, 10, 14, 58, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_access_level_sheet(ws, all_users):
    """Write a simple sheet listing all users with their Access Level (Admin/Standard/Sub-Standard)."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "IAM Users – Access Level Classification"
    c.font = Font(bold=True, color=COLORS["white"], size=13, name="Arial")
    c.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(["#", "User Name", "Access Level"], 1):
        hdr_cell(ws, 2, col, h)

    level_colors = {
        "Admin":        (COLORS["red_dark"],    COLORS["red_light"]),
        "Standard":     (COLORS["blue_dark"],   COLORS["blue_light"]),
        "Sub-Standard": (COLORS["grey_dark"],   COLORS["grey_light"]),
    }

    for idx, u in enumerate(all_users, 1):
        row = idx + 2
        cat = u["category"]
        if cat == "Admin":
            level = "Admin"
        elif cat in ("Privileged", "Power User", "Standard"):
            level = "Standard"
        else:
            level = "Sub-Standard"
        fg, bg = level_colors[level]
        alt = bg if idx % 2 == 0 else COLORS["white"]
        data_cell(ws, row, 1, idx, alt, align="center")
        data_cell(ws, row, 2, u["name"], alt)
        c = ws.cell(row=row, column=3, value=level)
        c.font = Font(name="Arial", size=10, bold=True, color=fg)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="top")
        c.border = THIN_BORDER

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.auto_filter.ref = f"A2:C{len(all_users) + 2}"


def main():
    global iam
    parser = argparse.ArgumentParser(description="IAM User Audit Report")
    parser.add_argument("--role-arn", help="IAM role ARN to assume")
    parser.add_argument("-o", "--output", default="iam_audit_report.xlsx", help="Output file")
    args = parser.parse_args()
    iam = get_iam_client(args.role_arn)

    print("Fetching IAM users...")
    raw_users = []
    paginator = iam.get_paginator("list_users")
    for page in paginator.paginate():
        raw_users.extend(page["Users"])
    all_users = []

    for i, user in enumerate(raw_users, 1):
        name = user["UserName"]
        print(f"  [{i}/{len(raw_users)}] {name}")
        detail = get_user_detail(name)
        all_users.append({
            "name": name,
            "created": user["CreateDate"].strftime("%Y-%m-%d"),
            **detail,
        })

    # Sort: admin first, then privileged, then rest
    order = {"Admin": 0, "Privileged": 1, "Power User": 2, "Standard": 3, "No Policies": 4}
    all_users.sort(key=lambda u: order.get(u["category"], 9))

    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "📊 Summary"
    write_summary_sheet(ws_summary, all_users)

    # Sheet 2: All Users
    ws_all = wb.create_sheet("👥 All Users")
    write_user_sheet(ws_all, all_users, "All IAM Users – Complete Audit")

    # Sheet 3: Admin users only
    admin_users = [u for u in all_users if u["category"] == "Admin"]
    ws_admin = wb.create_sheet("🔴 Admin Users")
    write_user_sheet(ws_admin, admin_users, "Admin-Level Users – Requires Immediate Review")

    # Sheet 4: Privileged/Power users
    priv_users = [u for u in all_users if u["category"] in ("Privileged", "Power User")]
    ws_priv = wb.create_sheet("🟠 Privileged Users")
    write_user_sheet(ws_priv, priv_users, "Privileged & Power Users")

    # Sheet 5: Standard users
    std_users = [u for u in all_users if u["category"] == "Standard"]
    ws_std = wb.create_sheet("🟢 Standard Users")
    write_user_sheet(ws_std, std_users, "Standard Users")

    # Sheet 6: Standard sub-category detail
    ws_sub = wb.create_sheet("📋 Standard Sub-Categories")
    write_sub_category_sheet(ws_sub, std_users)

    # Sheet 7: Access Level (Admin / Standard / Sub-Standard)
    ws_level = wb.create_sheet("🔑 Access Level")
    write_access_level_sheet(ws_level, all_users)

    wb.save(args.output)
    print(f"\n✅ Report saved: {args.output}")
    print(f"   Admin:      {len(admin_users)}")
    print(f"   Privileged: {len(priv_users)}")
    print(f"   Standard:   {len(std_users)}")
    print(f"   Total:      {len(all_users)}")


if __name__ == "__main__":
    main()
