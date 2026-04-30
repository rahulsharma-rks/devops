"""Generate a Current vs Required access report for all IAM users using Service Last Accessed data."""

import argparse
import time
import boto3
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = {
    "hdr": "1F3864", "white": "FFFFFF", "red": "C00000", "red_lt": "FFCCCC",
    "orange": "E36C09", "orange_lt": "FCE4D6", "green": "375623", "green_lt": "E2EFDA",
    "blue": "2F5496", "blue_lt": "D9E1F2", "grey": "595959", "grey_lt": "F2F2F2",
}
BORDER = Border(*(Side(style="thin", color="BFBFBF") for _ in range(4)))


def hdr(ws, row, col, val, bg=C["hdr"]):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color=C["white"], size=10, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


def cell(ws, row, col, val, bg=C["white"], bold=False, wrap=False, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=10, bold=bold)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = BORDER


def get_iam_client(role_arn=None):
    if role_arn:
        sts = boto3.client("sts")
        creds = sts.assume_role(RoleArn=role_arn, RoleSessionName="iam-access-review")["Credentials"]
        return boto3.client("iam", aws_access_key_id=creds["AccessKeyId"],
                            aws_secret_access_key=creds["SecretAccessKey"],
                            aws_session_token=creds["SessionToken"])
    return boto3.client("iam")


def get_all_users(iam):
    users = []
    for page in iam.get_paginator("list_users").paginate():
        users.extend(page["Users"])
    return users


def get_user_policies(iam, username):
    policies = []
    for page in iam.get_paginator("list_attached_user_policies").paginate(UserName=username):
        for p in page["AttachedPolicies"]:
            policies.append(f"[Direct] {p['PolicyName']}")
    for page in iam.get_paginator("list_user_policies").paginate(UserName=username):
        for name in page["PolicyNames"]:
            policies.append(f"[Inline] {name}")
    for page in iam.get_paginator("list_groups_for_user").paginate(UserName=username):
        for g in page["Groups"]:
            gn = g["GroupName"]
            for gp in iam.get_paginator("list_attached_group_policies").paginate(GroupName=gn):
                for p in gp["AttachedPolicies"]:
                    policies.append(f"[Group:{gn}] {p['PolicyName']}")
            for gp in iam.get_paginator("list_group_policies").paginate(GroupName=gn):
                for name in gp["PolicyNames"]:
                    policies.append(f"[Group:{gn}/Inline] {name}")
    return policies


def get_service_last_accessed(iam, user_arn):
    """Generate and retrieve service last accessed details for a user."""
    job_id = iam.generate_service_last_accessed_details(Arn=user_arn)["JobId"]
    for _ in range(30):
        resp = iam.get_service_last_accessed_details(JobId=job_id)
        if resp["JobStatus"] == "COMPLETED":
            return resp["ServicesLastAccessed"]
        if resp["JobStatus"] == "FAILED":
            return []
        time.sleep(1)
    return []


def classify_service(svc):
    """Classify a service entry as used, unused, or never accessed."""
    last = svc.get("LastAuthenticated") or svc.get("LastAuthenticatedEntity")
    if svc.get("LastAuthenticated"):
        days = (datetime.now(timezone.utc) - svc["LastAuthenticated"]).days
        return "used", days
    return "unused", None


def recommend(used_services, unused_services, total):
    if not unused_services:
        return "✅ Access matches usage — no changes needed"
    pct_unused = len(unused_services) / total * 100 if total else 0
    if pct_unused > 75:
        return f"🔴 {pct_unused:.0f}% of permitted services unused — strongly recommend generating least-privilege policy via Access Analyzer"
    if pct_unused > 40:
        return f"🟠 {pct_unused:.0f}% of permitted services unused — review and remove unnecessary service access"
    return f"🟡 {pct_unused:.0f}% of permitted services unused — minor cleanup recommended"


def write_detail_sheet(ws, users_data):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"IAM Current vs Required Access Review  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(bold=True, color=C["white"], size=13, name="Arial")
    c.fill = PatternFill("solid", fgColor=C["hdr"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "User Name", "Current Policies", "Services Permitted",
               "Services Used (Last 365d)", "Services Unused", "Used Count",
               "Unused Count", "Usage %", "Recommendation"]
    for col, h in enumerate(headers, 1):
        hdr(ws, 2, col, h)

    for idx, u in enumerate(users_data, 1):
        row = idx + 2
        pct = u["usage_pct"]
        if pct >= 80:
            bg = C["green_lt"]
        elif pct >= 50:
            bg = C["white"]
        elif pct >= 20:
            bg = C["orange_lt"]
        else:
            bg = C["red_lt"]
        alt = C["grey_lt"] if idx % 2 == 0 else bg

        cell(ws, row, 1, idx, alt, align="center")
        cell(ws, row, 2, u["name"], alt, bold=True)
        cell(ws, row, 3, "\n".join(u["policies"]) if u["policies"] else "—", alt, wrap=True)
        cell(ws, row, 4, u["permitted_count"], alt, align="center")
        cell(ws, row, 5, "\n".join(u["used_services"]) if u["used_services"] else "None", alt, wrap=True)
        cell(ws, row, 6, "\n".join(u["unused_services"]) if u["unused_services"] else "None", alt, wrap=True)
        cell(ws, row, 7, len(u["used_services"]), alt, align="center")
        cell(ws, row, 8, len(u["unused_services"]), alt, align="center")

        pc = ws.cell(row=row, column=9, value=pct / 100)
        pc.number_format = "0%"
        pc.font = Font(name="Arial", size=10, bold=True)
        pc.fill = PatternFill("solid", fgColor=alt)
        pc.alignment = Alignment(horizontal="center", vertical="top")
        pc.border = BORDER

        cell(ws, row, 10, u["recommendation"], alt, wrap=True)
        lines = max(len(u["used_services"]), len(u["unused_services"]),
                    len(u["policies"]), 1)
        ws.row_dimensions[row].height = max(20, min(lines * 15, 150))

    widths = [5, 28, 50, 12, 40, 40, 10, 10, 10, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A2:J{len(users_data) + 2}"


def write_summary_sheet(ws, users_data):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Current vs Required — Summary"
    c.font = Font(bold=True, color=C["white"], size=13, name="Arial")
    c.fill = PatternFill("solid", fgColor=C["hdr"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    total = len(users_data)
    over_provisioned = sum(1 for u in users_data if u["usage_pct"] < 50)
    well_sized = sum(1 for u in users_data if u["usage_pct"] >= 80)
    avg_usage = sum(u["usage_pct"] for u in users_data) / total if total else 0

    # KPIs
    kpis = [
        ("Total Users", total, C["blue"], C["blue_lt"]),
        ("Over-Provisioned (<50%)", over_provisioned, C["red"], C["red_lt"]),
        ("Well-Sized (≥80%)", well_sized, C["green"], C["green_lt"]),
        ("Avg Usage %", f"{avg_usage:.0f}%", C["orange"], C["orange_lt"]),
    ]
    for i, (label, val, fg, bg) in enumerate(kpis, 1):
        lc = ws.cell(row=3, column=i, value=label)
        lc.font = Font(bold=True, name="Arial", size=10, color=fg)
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = BORDER
        vc = ws.cell(row=4, column=i, value=val)
        vc.font = Font(bold=True, name="Arial", size=18, color=fg)
        vc.fill = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = BORDER
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 36
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Usage distribution
    row = 6
    ws.merge_cells(f"A{row}:E{row}")
    h = ws[f"A{row}"]
    h.value = "Usage Distribution"
    h.font = Font(bold=True, color=C["white"], size=11, name="Arial")
    h.fill = PatternFill("solid", fgColor=C["hdr"])
    h.alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for col, hd in enumerate(["Usage Band", "Count", "% of Users", "Risk", "Action"], 1):
        hdr(ws, row, col, hd)

    bands = [
        ("0–20% usage",  lambda u: u["usage_pct"] < 20,  C["red_lt"],    "🔴 Critical", "Generate least-privilege policy via Access Analyzer; likely needs full policy replacement"),
        ("20–50% usage", lambda u: 20 <= u["usage_pct"] < 50, C["orange_lt"], "🟠 High", "Remove unused service permissions; review with team lead"),
        ("50–80% usage", lambda u: 50 <= u["usage_pct"] < 80, C["white"],     "🟡 Medium", "Minor cleanup; remove clearly unused services"),
        ("80–100% usage",lambda u: u["usage_pct"] >= 80, C["green_lt"],  "🟢 Low",     "Well-sized; routine review only"),
    ]
    for band_label, fn, bg, risk, action in bands:
        row += 1
        cnt = sum(1 for u in users_data if fn(u))
        cell(ws, row, 1, band_label, bg, bold=True)
        cell(ws, row, 2, cnt, bg, align="center")
        pc = ws.cell(row=row, column=3, value=cnt / total if total else 0)
        pc.number_format = "0.0%"
        pc.font = Font(name="Arial", size=10)
        pc.fill = PatternFill("solid", fgColor=bg)
        pc.alignment = Alignment(horizontal="center", vertical="top")
        pc.border = BORDER
        cell(ws, row, 4, risk, bg)
        cell(ws, row, 5, action, bg, wrap=True)
        ws.row_dimensions[row].height = 30

    # Top over-provisioned users
    over = sorted(users_data, key=lambda u: u["usage_pct"])[:20]
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    h = ws[f"A{row}"]
    h.value = "⚠ Most Over-Provisioned Users"
    h.font = Font(bold=True, color=C["white"], size=11, name="Arial")
    h.fill = PatternFill("solid", fgColor=C["red"])
    h.alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for col, hd in enumerate(["User Name", "Permitted", "Used", "Unused", "Usage %"], 1):
        hdr(ws, row, col, hd, bg=C["red"])

    for u in over:
        row += 1
        bg = C["red_lt"] if row % 2 == 0 else C["white"]
        cell(ws, row, 1, u["name"], bg, bold=True)
        cell(ws, row, 2, u["permitted_count"], bg, align="center")
        cell(ws, row, 3, len(u["used_services"]), bg, align="center")
        cell(ws, row, 4, len(u["unused_services"]), bg, align="center")
        pc = ws.cell(row=row, column=5, value=u["usage_pct"] / 100)
        pc.number_format = "0%"
        pc.font = Font(name="Arial", size=10, bold=True,
                       color=C["red"] if u["usage_pct"] < 20 else C["orange"])
        pc.fill = PatternFill("solid", fgColor=bg)
        pc.alignment = Alignment(horizontal="center", vertical="top")
        pc.border = BORDER

    ws.column_dimensions["E"].width = 50


def main():
    parser = argparse.ArgumentParser(description="IAM Current vs Required Access Review")
    parser.add_argument("--role-arn", help="IAM role ARN to assume")
    parser.add_argument("-o", "--output", default="iam_access_review.xlsx", help="Output file")
    args = parser.parse_args()

    iam = get_iam_client(args.role_arn)

    print("Fetching IAM users...")
    raw_users = get_all_users(iam)
    print(f"Found {len(raw_users)} users. Generating service last accessed data...\n")

    users_data = []
    for i, user in enumerate(raw_users, 1):
        name = user["UserName"]
        arn = user["Arn"]
        print(f"  [{i}/{len(raw_users)}] {name}...", end=" ", flush=True)

        policies = get_user_policies(iam, name)
        services = get_service_last_accessed(iam, arn)

        used, unused = [], []
        for svc in services:
            svc_name = svc["ServiceName"]
            status, days = classify_service(svc)
            if status == "used" and days is not None and days <= 365:
                used.append(f"{svc_name} ({days}d ago)")
            else:
                unused.append(svc_name)

        permitted = len(services)
        used_count = len(used)
        pct = (used_count / permitted * 100) if permitted else 100

        users_data.append({
            "name": name,
            "policies": policies,
            "permitted_count": permitted,
            "used_services": used,
            "unused_services": unused,
            "usage_pct": round(pct, 1),
            "recommendation": recommend(used, unused, permitted),
        })
        print(f"✓ {used_count}/{permitted} services used ({pct:.0f}%)")

    # Sort by usage % ascending (most over-provisioned first)
    users_data.sort(key=lambda u: u["usage_pct"])

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "📊 Summary"
    write_summary_sheet(ws_summary, users_data)

    ws_detail = wb.create_sheet("📋 Detail")
    write_detail_sheet(ws_detail, users_data)

    wb.save(args.output)
    over = sum(1 for u in users_data if u["usage_pct"] < 50)
    print(f"\n✅ Report saved: {args.output}")
    print(f"   Total users:       {len(users_data)}")
    print(f"   Over-provisioned:  {over}")
    print(f"   Avg usage:         {sum(u['usage_pct'] for u in users_data) / len(users_data):.0f}%")


if __name__ == "__main__":
    main()
